#! /usr/bin/env python3

# ladiesEclectic.py - Updates a specific excel spreadsheet from the command line

# path to file on Windows = 'S:\Golf\TOURNAMENTS\Tournaments_Clinics\ALL EVENTS\
# 2 Women's Golf Events\Ladies Eclectic\2018-2019\Eclectic Scores 2018-2019.xlsx'
# Windows shebang line '#! python3'

import openpyxl, datetime, sys
from openpyxl import load_workbook

def saveCopyWorkbook():                           
    """Save copy of older workbook"""
    now = datetime.datetime.now()
    fileBasename = now.strftime('%m_%d_%Y')
    saveFile = 'practiceEclecticScores' + fileBasename + '.xlsx'
    wb.save(saveFile)
    print('Copy of current document saved as %s' % (saveFile))

def saveWorkbook():                               
    """Save over current workbook"""
    workbookName = 'practiceEclecticScores.xlsx'
    wb.save(workbookName)
    print('Document saved as %s' % (workbookName))

def compareScoresSouth(rowNum):                   
    """Change eclectic score if inputted score is lower"""
    global scores
    scoresNotChanged = 0
    scoresChanged = 0
    scoringRowSouth = rowNum + 6                  # row based on course
    for i in range(len(scores)):
        assert int(scores[i]) < 20, 'Make sure scores are entered with a space inbetween.'
        if int(scores[i]) < int(ws.cell(row=scoringRowSouth, column=2+i).value):
            ws.cell(row=scoringRowSouth, column=2+i).value = int(scores[i])
            scoresChanged += 1
        else:
            scoresNotChanged += 1
    print('There were ' + str(scoresNotChanged) + ' scores that were not changed')       
    print('There were ' + str(scoresChanged) + ' scores that were improved!')        

def compareScoresNorth(rowNum):                         
    """Change eclectic score if inputted score is lower"""
    global scores
    scoresNotChanged = 0
    scoresChanged = 0
    scoringRowNorth = rowNum + 12                 # row based on course
    for i in range(len(scores)):
        assert int(scores[i]) < 20, 'Make sure scores are entered with a space inbetween.'
        if int(scores[i]) < int(ws.cell(row=scoringRowNorth, column=2+i).value):
            ws.cell(row=scoringRowNorth, column=2+i).value = int(scores[i])
            scoresChanged += 1
        else:
            scoresNotChanged += 1
    print('There were ' + str(scoresChanged) + ' scores that were improved!')
    print('There were ' + str(scoresNotChanged) + ' scores that were not changed')

def name_Check(name):                             
    """Check if name is found in workbook value"""
    for rowNum in range(1, ws.max_row):
        golferName = ws.cell(row=rowNum, column=6).value
        if str(golferName) != name.upper():
            trueAndRowNum = (False, rowNum)
        elif str(golferName) == name.upper():
            trueAndRowNum = (True, rowNum)
            return trueAndRowNum
            break

def check_Course(course):
    """Check if course variable input is usable"""
    if course.upper() == 'N':
        return True
    elif course.upper() == 'S':
        return True
    else:
        return False


# Create command line arguments for inputs
name = str(sys.argv[1] + ', ' + sys.argv[2])      # case insensitive
course = str(sys.argv[3])
scores = sys.argv[4:]

#  Open the latest version of ladiesPracticeEclecticScores
wb = load_workbook('practiceEclecticScores.xlsx')  # loads current workbook before any edits
saveCopyWorkbook()                 

wb = load_workbook('practiceEclecticScores.xlsx')  # loads current workbook for editing
ws = wb['nEW17_18']


try:
    returnValue = name_Check(name)                 # sees if name is in document
    nameCheck = returnValue[0]                     # gets boolean from return value
    checkCourse = check_Course(course)             # sees if course is valid input
    if nameCheck == True and checkCourse == True:  # check that name and course are usable
       rowNum = returnValue[1]                     # gets row number from wb that ladies name is in
       if course.upper() == 'S':
           compareScoresSouth(rowNum)
           saveWorkbook()
       else:
           compareScoresNorth(rowNum)
           saveWorkbook()
    else:
        print('Something is wrong.')   
except TypeError:
    print('An error occurred. Please check spelling, course, or scores.')
    wb.close()


